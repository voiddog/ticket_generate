import datetime
import os
import bisect
import random

from openpyxl import load_workbook
from template import *


def read_menu_list(xlsx_file_path) -> list[MenuItem]:
    wb = load_workbook(xlsx_file_path)
    sheet = wb.active
    ret = []
    for row in sheet.iter_rows(min_row=2):
        ret.append(MenuItem(
            row[0].value,
            1,
            row[1].value
        ))
    return ret


def generate_random_menu_list(menu_list: list[MenuItem], count) -> list[MenuItem]:
    if count == len(menu_list):
        return menu_list

    ret = []
    while count > 0:
        index = random.randint(0, len(menu_list) - 1)
        while True:
            item = menu_list[index]
            if item in ret:
                index = (index + 1) % len(menu_list)
            else:
                ret.append(item)
                break
        count = count - 1
    return ret


def generate_times(
        date: datetime.date,
        order_prop_list: dict[datetime.time, float],
        start=datetime.time(hour=8),
        end=datetime.time(hour=22),
        min_interval=datetime.timedelta(minutes=3),
        max_interval=datetime.timedelta(minutes=10),
) -> list[datetime.datetime]:
    if not order_prop_list or len(order_prop_list) == 0:
        raise "订单生成概率为 None"
    t = start
    keys = []
    keys.extend(order_prop_list.keys())
    sorted(keys)
    ret = []
    while t < end:
        k_index = bisect.bisect_left(keys, t)
        if k_index >= len(keys):
            k_index = len(keys) - 1
        p = order_prop_list[keys[k_index]]
        if random.random() < p:
            ret.append(datetime.datetime(date.year, date.month, date.day, t.hour, t.minute, t.second))
        second_delta = random.randint(min_interval.seconds, max_interval.seconds)
        new_second = t.hour * 3600 + t.minute * 60 + t.second + second_delta
        t = datetime.time(new_second // 3600, (new_second % 3600) // 60, new_second % 60)
    return ret


def render_template_a(start_day: datetime.date, end_day: datetime.date):
    template = Template(
        'template_a.xlsx',
        RenderRange((4, 1), render_format="单号:%s"),
        RenderRange((4, 2), render_format="台牌:%s"),
        RenderRange((5, 1), render_format="开单:%s"),
        RenderRange((5, 2), render_format="收银:%s"),
        8,
        key_range=RenderRange((10, 1), (10, 3), "key:%s"),
        print_time_range=RenderRange((11, 1), (11, 3), "打印时间:%s"),
    )
    # 生成菜单
    menu_list = read_menu_list('menu.xlsx')

    # 生成订单日期数据
    date_list = []
    day = start_day
    while day <= end_day:
        date_list.extend(generate_times(day, {
            # 上午时间段
            datetime.time(hour=10): 0.0,
            datetime.time(hour=11): 0.3,
            datetime.time(hour=12): 0.5,
            datetime.time(hour=13): 0.2,
            datetime.time(hour=14): 0.1,
            datetime.time(hour=15): 0.0,
        }))
        day = day + datetime.timedelta(days=1)

    # 渲染菜单
    for date_item in date_list:
        wb = template.render(TicketData(
            '0075',
            '20',
            '01|王明浩',
            '01|王明浩',
            generate_random_menu_list(menu_list, random.randint(2, 10)),
            date_item
        ))

        file_path = "out/%s.xlsx" % date_item.strftime("%Y-%m-%d-%H-%M-%S")
        if os.path.exists(file_path):
            os.remove(file_path)
        wb.save(file_path)
        wb = xw.Book(file_path)
        wb.close()


if __name__ == '__main__':
    render_template_a(
        datetime.date(2021, 6, 1),
        datetime.date(2021, 6, 30))
